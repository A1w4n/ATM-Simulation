import hashlib
import os
import openpyxl
from datetime import datetime

# ─────────────────────────────────────────────
#  Path configuration  – edit these two lines
# ─────────────────────────────────────────────
BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE_DIR, "atmData.xlsx")


# ── Helpers ──────────────────────────────────

def hash_pin(pin: str) -> str:
    """Return the SHA-256 hex digest of a PIN string."""
    return hashlib.sha256(pin.encode("utf-8")).hexdigest()


def _load_workbook():
    """Open the workbook and return (wb, sheet).  Raises FileNotFoundError if missing."""
    if not os.path.exists(XLSX_PATH):
        raise FileNotFoundError(
            f"atmData.xlsx not found at:\n{XLSX_PATH}\n\n"
            "Please update BASE_DIR / XLSX_PATH inside ATM_auth.py."
        )
    wb = openpyxl.load_workbook(XLSX_PATH)
    return wb, wb.active


# ── Account lookup ────────────────────────────

def get_account(account_number: int) -> dict | None:
    """
    Look up *account_number* in the workbook.

    Returns a dict with keys  accountNumber, pin, balance
    or None if the account does not exist.

    Expected column layout (1-based):
        A  accountNumber  (int)
        B  hashed PIN     (str, SHA-256 hex)
        C  balance        (float)
    """
    try:
        wb, sheet = _load_workbook()
    except FileNotFoundError:
        raise

    for row in sheet.iter_rows(values_only=True):
        if row[0] == account_number:
            return {
                "accountNumber": row[0],
                "pin":           row[1],
                "balance":       float(row[2]),
            }
    return None


def verify_pin(account_number: int, pin: str) -> dict | None:
    """
    Fetch the account and verify the PIN.

    Returns the account dict on success, None on failure.
    Raises FileNotFoundError if the workbook cannot be found.
    """
    account = get_account(account_number)
    if account and hash_pin(pin) == account["pin"]:
        return account
    return None


# ── Balance mutations ─────────────────────────

def update_balance(account_number: int, new_balance: float) -> None:
    """Persist *new_balance* for *account_number* back to the workbook."""
    wb, sheet = _load_workbook()
    for row in sheet.iter_rows():
        if row[0].value == account_number:
            row[2].value = round(new_balance, 2)
            break
    wb.save(XLSX_PATH)


def deposit(account_number: int, amount: float, current_balance: float) -> float:
    """Add *amount* to the balance, persist, and return the new balance."""
    if amount <= 0:
        raise ValueError("Deposit amount must be positive.")
    new_balance = round(current_balance + amount, 2)
    update_balance(account_number, new_balance)
    record_transaction(account_number, "Deposit", amount=amount, balance=new_balance)
    return new_balance


def withdraw(account_number: int, amount: float, current_balance: float,
             service_fee: float = 18.0) -> float:
    """
    Deduct *amount* + *service_fee* from the balance, persist, and return the new balance.
    Raises ValueError if funds are insufficient.
    """
    if amount <= 0:
        raise ValueError("Withdrawal amount must be positive.")
    total = amount + service_fee
    if current_balance < total:
        raise ValueError(
            f"Insufficient funds.  "
            f"You need Php {total:.2f} (amount + Php {service_fee:.2f} fee) "
            f"but your balance is only Php {current_balance:.2f}."
        )
    new_balance = round(current_balance - total, 2)
    update_balance(account_number, new_balance)
    record_transaction(account_number, "Withdraw", amount=amount, balance=new_balance)
    return new_balance


# ── Transaction logging ───────────────────────

def _record_path(account_number: int) -> str:
    return os.path.join(BASE_DIR, f"{account_number}_records.txt")


def record_transaction(account_number: int, action: str,
                       amount: float | None = None,
                       balance: float | None = None) -> None:
    """Append a timestamped line to the account's transaction log."""
    timestamp = datetime.now().strftime("%m-%d-%Y %H:%M:%S")
    line = f"[{timestamp}] Action: {action}"
    if amount  is not None: line += f", Amount: Php {amount:.2f}"
    if balance is not None: line += f", New Balance: Php {balance:.2f}"
    line += "\n"
    with open(_record_path(account_number), "a") as f:
        f.write(line)


def get_transaction_history(account_number: int) -> str:
    """Return the full text of the transaction log, or an empty string if none exists."""
    path = _record_path(account_number)
    if os.path.exists(path):
        with open(path, "r") as f:
            return f.read()
    return ""


def build_receipt(account_number: int, action: str,
                  amount: float | None = None,
                  balance: float | None = None) -> str:
    """Return a formatted receipt string."""
    timestamp = datetime.now().strftime("%m-%d-%Y %H:%M:%S")
    lines = [
        "╔══════════════════════════════════════╗",
        "║       TRANSACTION RECEIPT            ║",
        "╠══════════════════════════════════════╣",
        f"  Account  : {account_number}",
        f"  Date     : {timestamp}",
        f"  Action   : {action}",
    ]
    if amount  is not None: lines.append(f"  Amount   : Php {amount:.2f}")
    if balance is not None: lines.append(f"  Balance  : Php {balance:.2f}")
    lines.append("╚══════════════════════════════════════╝")
    return "\n".join(lines)
